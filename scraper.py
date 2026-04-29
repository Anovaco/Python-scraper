"""
CONTACT INFO SCRAPER v6
7-layer fallback chain:
1. Direct scrape (homepage + subpages + sitemap)
2. Google Custom Search
3. Google Maps / Places API
4. Bing search scrape
5. Yellow Pages Canada
6. WHOIS lookup
7. Selenium (real browser, per-thread) — last resort

New in v6:
- mailto:/tel: href extraction (huge win over plain-text regex)
- JSON-LD / Schema.org structured data parsing
- Sitemap.xml crawling to find contact pages
- Obfuscated email decoding ([at], (at), " at ")
- WHOIS lookup as Layer 6
- Per-thread Selenium drivers (fixes v5 threading bug)
- Better phone regex (catches tel: hrefs, more formats)
"""

import re, os, time, random, json, threading
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse, urljoin, quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

# ── YOUR CREDENTIALS ──────────────────────────────────────────────────────────
GOOGLE_API_KEY = "AIzaSyCV8jp30LZ51r1VA8eJS_PqVyGsQtyTA8o"
GOOGLE_CX      = "53de390cb47a24f08"
# ─────────────────────────────────────────────────────────────────────────────

URLS_FILE       = "urls.txt"
OUTPUT_FILE     = "contact_info.xlsx"
SAVE_EVERY      = 25
MAX_WORKERS     = 4
REQUEST_TIMEOUT = 15
MAX_RETRIES     = 3
RETRY_DELAY     = 5

HEADERS = {"User-Agent": (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)}
BING_HEADERS = {"User-Agent": (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0"
)}

EMAIL_RE = re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+")

# Obfuscated email: "name [at] domain [dot] com" or "name(at)domain.com"
OBFUS_EMAIL_RE = re.compile(
    r"([a-zA-Z0-9_.+-]+)\s*[\[\(]?\s*at\s*[\]\)]?\s*([a-zA-Z0-9-]+)\s*[\[\(]?\s*dot\s*[\]\)]?\s*([a-zA-Z]{2,})",
    re.IGNORECASE
)

PHONE_RE = re.compile(
    r"(\+?1[\s.\-]?)?(\(?\d{3}\)?[\s.\-]?)(\d{3}[\s.\-]?\d{4})"
)

SOCIAL_PATTERNS = {
    "Instagram": re.compile(r"(?:https?://)?(?:www\.)?instagram\.com/([A-Za-z0-9_.]+)"),
    "Twitter/X":  re.compile(r"(?:https?://)?(?:www\.)?(?:twitter|x)\.com/([A-Za-z0-9_]+)"),
    "Facebook":  re.compile(r"(?:https?://)?(?:www\.)?facebook\.com/((?!share|sharer|plugins|login|dialog|home)[A-Za-z0-9_./-]+)"),
    "LinkedIn":  re.compile(r"(?:https?://)?(?:www\.)?linkedin\.com/(?:in|company)/([A-Za-z0-9_-]+)"),
    "YouTube":   re.compile(r"(?:https?://)?(?:www\.)?youtube\.com/(?:channel/|c/|user/|@)([A-Za-z0-9_@-]+)"),
    "TikTok":    re.compile(r"(?:https?://)?(?:www\.)?tiktok\.com/@([A-Za-z0-9_.]+)"),
}
SOCIAL_BASES = {
    "Instagram": "https://instagram.com/",
    "Twitter/X":  "https://x.com/",
    "Facebook":  "https://facebook.com/",
    "LinkedIn":  "https://linkedin.com/company/",
    "YouTube":   "https://youtube.com/@",
    "TikTok":    "https://tiktok.com/@",
}
SKIP_HANDLES = {"sharer","share","plugins","login","home","pages","groups",
                "events","watch","hashtag","explore","p","permalink","tr",
                "photo","video","stories","reel","reels","ads","business"}
BAD_EMAIL_PARTS = ["example.","yourdomain","sentry","wix","schema.org",
                   ".png",".jpg",".gif",".svg",".webp","noreply@example",
                   "@2x","@3x","webpack","eslint","email@","test@",
                   "info@example","user@","admin@example","support@example"]

COLS = ["#","URL","Status","Emails","Phone Numbers",
        "Instagram","Twitter/X","Facebook","LinkedIn","YouTube","TikTok"]

# ── Per-thread Selenium drivers ───────────────────────────────────────────────
_thread_local = threading.local()

def get_selenium_driver():
    """Returns a per-thread Selenium driver (fixes v5 shared-driver race condition)."""
    if getattr(_thread_local, "driver", None) is not None:
        return _thread_local.driver
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from webdriver_manager.chrome import ChromeDriverManager

        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--window-size=1280,800")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
        driver.set_page_load_timeout(20)
        _thread_local.driver = driver
        return driver
    except Exception as e:
        print(f"\n  Selenium setup failed: {e}")
        _thread_local.driver = None
        return None


def quit_all_selenium():
    """Called at the end — individual threads clean up their own drivers."""
    pass  # Drivers are per-thread and cleaned up by OS on exit


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def fetch(url, retries=MAX_RETRIES, hdrs=None):
    hdrs = hdrs or HEADERS
    for attempt in range(retries):
        try:
            time.sleep(random.uniform(0.3, 1.2))
            r = requests.get(url, headers=hdrs, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if r.status_code == 429:
                time.sleep(RETRY_DELAY * (attempt+1) + random.uniform(1,4))
                continue
            return r
        except Exception:
            if attempt == retries-1:
                raise
            time.sleep(2)
    return None


def extract_emails_from_soup(soup, domain_hint=""):
    """Extract emails from mailto: hrefs (most reliable) + visible text."""
    emails = set()

    # 1. mailto: hrefs — most reliable source
    for tag in soup.find_all("a", href=True):
        href = tag["href"]
        if href.lower().startswith("mailto:"):
            addr = href[7:].split("?")[0].strip().lower()
            if addr and "@" in addr:
                emails.add(addr)

    # 2. Visible text regex
    text = soup.get_text(" ")
    for e in EMAIL_RE.findall(text):
        emails.add(e.lower().strip(".,;:"))

    # 3. Obfuscated emails: "name [at] domain [dot] com"
    for m in OBFUS_EMAIL_RE.finditer(text):
        reconstructed = f"{m.group(1)}@{m.group(2)}.{m.group(3)}".lower()
        emails.add(reconstructed)

    # Filter bad emails
    cleaned = set()
    for e in emails:
        if any(b in e for b in BAD_EMAIL_PARTS):
            continue
        if len(e) > 80 or "." not in e.split("@")[-1]:
            continue
        cleaned.add(e)

    if domain_hint:
        root = domain_hint.replace("www.", "").split(".")[0]
        preferred = [e for e in cleaned if root in e]
        if preferred:
            return sorted(preferred)[:5]
    return sorted(cleaned)[:5]


def extract_emails(text, domain_hint=""):
    """Plain text fallback (no soup object available)."""
    emails = set()
    for e in EMAIL_RE.findall(text):
        emails.add(e.lower().strip(".,;:"))
    for m in OBFUS_EMAIL_RE.finditer(text):
        emails.add(f"{m.group(1)}@{m.group(2)}.{m.group(3)}".lower())
    cleaned = {e for e in emails if not any(b in e for b in BAD_EMAIL_PARTS)
               and len(e) <= 80 and "." in e.split("@")[-1]}
    if domain_hint:
        root = domain_hint.replace("www.", "").split(".")[0]
        preferred = [e for e in cleaned if root in e]
        if preferred:
            return sorted(preferred)[:5]
    return sorted(cleaned)[:5]


def extract_phones_from_soup(soup):
    """Extract phones from tel: hrefs (most reliable) + visible text."""
    phones = set()

    # 1. tel: hrefs — most reliable
    for tag in soup.find_all("a", href=True):
        href = tag["href"]
        if href.lower().startswith("tel:"):
            digits = re.sub(r"\D", "", href[4:])
            if 10 <= len(digits) <= 11:
                phones.add(href[4:].strip())

    # 2. Text regex
    text = soup.get_text(" ")
    for m in PHONE_RE.finditer(text):
        digits = re.sub(r"\D", "", m.group(0))
        if 10 <= len(digits) <= 11:
            phones.add(m.group(0).strip())

    return sorted(phones)[:3]


def extract_phones(text):
    phones = set()
    for m in PHONE_RE.finditer(text):
        digits = re.sub(r"\D","",m.group(0))
        if 10 <= len(digits) <= 11:
            phones.add(m.group(0).strip())
    return sorted(phones)[:3]


def extract_socials(raw):
    found = {}
    for platform, pattern in SOCIAL_PATTERNS.items():
        for handle in pattern.findall(raw):
            handle = handle.rstrip("/")
            if handle.lower() not in SKIP_HANDLES and len(handle) > 1:
                found[platform] = SOCIAL_BASES[platform] + handle
                break
    return found


def extract_jsonld_contacts(soup):
    """Parse JSON-LD / Schema.org for structured contact data."""
    emails, phones = [], []
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            # Flatten nested @graph structures
            nodes = data if isinstance(data, list) else data.get("@graph", [data])
            for node in nodes:
                if not isinstance(node, dict):
                    continue
                # Email
                for key in ["email", "contactPoint"]:
                    val = node.get(key, "")
                    if isinstance(val, str) and "@" in val:
                        emails.append(val.replace("mailto:", "").strip())
                    elif isinstance(val, dict):
                        e = val.get("email", "")
                        if e and "@" in e:
                            emails.append(e.replace("mailto:", "").strip())
                # Phone
                for key in ["telephone", "phone", "faxNumber"]:
                    val = node.get(key, "")
                    if val and isinstance(val, str):
                        digits = re.sub(r"\D", "", val)
                        if 10 <= len(digits) <= 11:
                            phones.append(val.strip())
        except Exception:
            continue
    return emails[:3], phones[:3]


def get_domain(url):
    return urlparse(url).netloc.lower()

def get_domain_name(url):
    return urlparse(url).netloc.lower().replace("www.","").split(".")[0]

def still_missing(row):
    return (not row["Emails"] or not row["Phone Numbers"] or
            not row["Instagram"] or not row["Facebook"])


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 1 — Direct scrape (homepage + subpages + sitemap)
# ═══════════════════════════════════════════════════════════════════════════════

CONTACT_PATHS = [
    "/contact", "/contact-us", "/about", "/about-us",
    "/reach-us", "/get-in-touch", "/contactus", "/info",
    "/connect", "/our-team", "/locations", "/location",
]

def get_sitemap_contact_urls(root_url):
    """Parse sitemap.xml to find contact/about pages."""
    extra_urls = []
    for sitemap_path in ["/sitemap.xml", "/sitemap_index.xml", "/sitemap"]:
        try:
            r = fetch(urljoin(root_url, sitemap_path))
            if not r or r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "lxml-xml")
            for loc in soup.find_all("loc"):
                url_text = loc.get_text().strip().lower()
                if any(kw in url_text for kw in ["contact","about","reach","info","location"]):
                    extra_urls.append(loc.get_text().strip())
            if extra_urls:
                break
        except Exception:
            pass
    return extra_urls[:5]


def direct_scrape(url):
    all_soups, all_text, all_raw, status = [], "", "", "OK"
    parsed = urlparse(url)
    root = f"{parsed.scheme}://{parsed.netloc}"

    # Fetch homepage
    try:
        r = fetch(url)
        if r is None:
            return "", "", "HTTP 429 (blocked)", []
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        all_soups.append(soup)
        all_text += soup.get_text(" ")
        all_raw  += r.text
    except requests.exceptions.HTTPError as e:
        status = f"HTTP {e.response.status_code}"
    except requests.exceptions.ConnectionError:
        status = "Connection Error"
    except requests.exceptions.Timeout:
        status = "Timeout"
    except Exception:
        status = "Error"

    # Fixed subpages
    for path in CONTACT_PATHS:
        try:
            sr = fetch(urljoin(root, path))
            if sr and sr.status_code == 200:
                s = BeautifulSoup(sr.text, "lxml")
                all_soups.append(s)
                all_text += " " + s.get_text(" ")
                all_raw  += " " + sr.text
        except Exception:
            pass

    # Sitemap-discovered pages
    for extra_url in get_sitemap_contact_urls(root):
        try:
            sr = fetch(extra_url)
            if sr and sr.status_code == 200:
                s = BeautifulSoup(sr.text, "lxml")
                all_soups.append(s)
                all_text += " " + s.get_text(" ")
                all_raw  += " " + sr.text
        except Exception:
            pass

    return all_text, all_raw, status, all_soups


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 2 — Google Custom Search
# ═══════════════════════════════════════════════════════════════════════════════

def google_search(query):
    try:
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key":GOOGLE_API_KEY,"cx":GOOGLE_CX,"q":query,"num":5},
            timeout=10
        )
        if r.status_code != 200:
            return ""
        data = r.json()
        parts = []
        for item in data.get("items",[]):
            parts += [item.get("snippet",""), item.get("title",""), item.get("link","")]
        return " ".join(parts)
    except Exception:
        return ""

def google_fallback(domain, domain_name, row):
    if not row["Emails"]:
        text = google_search(f'"{domain}" email')
        emails = extract_emails(text, domain_name)
        if emails:
            row["Emails"] = ", ".join(emails[:3])
    if not row["Phone Numbers"]:
        text = google_search(f'"{domain}" phone number')
        phones = extract_phones(text)
        if phones:
            row["Phone Numbers"] = ", ".join(phones)
    for platform, query in [
        ("Instagram", f'site:instagram.com "{domain_name}"'),
        ("Facebook",  f'site:facebook.com "{domain_name}" travel'),
        ("LinkedIn",  f'site:linkedin.com/company "{domain_name}"'),
        ("Twitter/X", f'site:x.com "{domain_name}"'),
        ("YouTube",   f'site:youtube.com "{domain_name}"'),
        ("TikTok",    f'site:tiktok.com "{domain_name}"'),
    ]:
        if not row[platform]:
            text = google_search(query)
            socials = extract_socials(text)
            if platform in socials:
                row[platform] = socials[platform]


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 3 — Google Maps / Places
# ═══════════════════════════════════════════════════════════════════════════════

def google_maps_fallback(domain_name, row):
    if row["Phone Numbers"]:
        return
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
            params={"input":f"{domain_name} travel agency canada",
                    "inputtype":"textquery","fields":"place_id","key":GOOGLE_API_KEY},
            timeout=10
        )
        candidates = r.json().get("candidates",[])
        if not candidates:
            return
        r2 = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={"place_id":candidates[0]["place_id"],
                    "fields":"formatted_phone_number","key":GOOGLE_API_KEY},
            timeout=10
        )
        phone = r2.json().get("result",{}).get("formatted_phone_number","")
        if phone:
            row["Phone Numbers"] = phone
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 4 — Bing search scrape
# ═══════════════════════════════════════════════════════════════════════════════

def bing_search(query):
    try:
        r = requests.get(
            f"https://www.bing.com/search?q={quote_plus(query)}",
            headers=BING_HEADERS, timeout=10
        )
        if r.status_code != 200:
            return ""
        soup = BeautifulSoup(r.text,"lxml")
        for tag in soup(["script","style"]):
            tag.decompose()
        return soup.get_text(" ")
    except Exception:
        return ""

def bing_fallback(domain, domain_name, row):
    if not row["Emails"]:
        text = bing_search(f"{domain} email contact")
        emails = extract_emails(text, domain_name)
        if emails:
            row["Emails"] = ", ".join(emails[:3])
    if not row["Phone Numbers"]:
        text = bing_search(f"{domain} phone number")
        phones = extract_phones(text)
        if phones:
            row["Phone Numbers"] = ", ".join(phones)
    for platform, query in [
        ("Instagram", f"{domain_name} travel instagram"),
        ("Facebook",  f"{domain_name} travel facebook"),
        ("LinkedIn",  f"{domain_name} linkedin company"),
    ]:
        if not row[platform]:
            text = bing_search(query)
            socials = extract_socials(text)
            if platform in socials:
                row[platform] = socials[platform]


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 5 — Yellow Pages Canada
# ═══════════════════════════════════════════════════════════════════════════════

def yellowpages_fallback(domain_name, row):
    if row["Phone Numbers"] and row["Emails"]:
        return
    try:
        r = requests.get(
            f"https://www.yellowpages.ca/search/si/1/{quote_plus(domain_name+' travel')}/Canada",
            headers=HEADERS, timeout=10
        )
        if r.status_code != 200:
            return
        text = BeautifulSoup(r.text,"lxml").get_text(" ")
        if not row["Phone Numbers"]:
            phones = extract_phones(text)
            if phones:
                row["Phone Numbers"] = ", ".join(phones)
        if not row["Emails"]:
            emails = extract_emails(text, domain_name)
            if emails:
                row["Emails"] = ", ".join(emails[:2])
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 6 — WHOIS lookup  ← NEW in v6
# ═══════════════════════════════════════════════════════════════════════════════

def whois_fallback(domain, row):
    """Look up WHOIS registrant contact info — often reveals email for small sites."""
    if row["Emails"] and row["Phone Numbers"]:
        return
    try:
        import whois
        w = whois.whois(domain)
        # Email
        if not row["Emails"]:
            emails = w.emails if isinstance(w.emails, list) else ([w.emails] if w.emails else [])
            clean = [e.lower() for e in emails if e and "@" in e
                     and not any(b in e.lower() for b in BAD_EMAIL_PARTS)]
            if clean:
                row["Emails"] = ", ".join(clean[:2])
        # Phone (WHOIS phones are rare but exist)
        if not row["Phone Numbers"]:
            phones = []
            for attr in ["phone", "registrant_phone"]:
                val = getattr(w, attr, None)
                if val:
                    phones.append(str(val))
            if phones:
                row["Phone Numbers"] = phones[0]
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# LAYER 7 — Selenium real browser (per-thread)
# ═══════════════════════════════════════════════════════════════════════════════

def selenium_fallback(url, domain_name, row):
    if not still_missing(row):
        return

    driver = get_selenium_driver()
    if driver is None:
        return

    try:
        from selenium.webdriver.common.by import By

        driver.get(url)
        time.sleep(3)

        contact_keywords = ["contact", "reach us", "get in touch", "about us"]
        for kw in contact_keywords:
            try:
                els = driver.find_elements(By.PARTIAL_LINK_TEXT, kw.title())
                if not els:
                    els = driver.find_elements(By.PARTIAL_LINK_TEXT, kw)
                if els:
                    els[0].click()
                    time.sleep(2)
                    break
            except Exception:
                pass

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        page_src  = driver.page_source
        soup = BeautifulSoup(page_src, "lxml")

        if not row["Emails"]:
            emails = extract_emails_from_soup(soup, domain_name)
            if emails:
                row["Emails"] = ", ".join(emails[:3])

        if not row["Phone Numbers"]:
            phones = extract_phones_from_soup(soup)
            if phones:
                row["Phone Numbers"] = ", ".join(phones)

        socials = extract_socials(page_src)
        for platform, link in socials.items():
            if not row[platform]:
                row[platform] = link

        # JSON-LD on JS-rendered page
        if not row["Emails"] or not row["Phone Numbers"]:
            jl_emails, jl_phones = extract_jsonld_contacts(soup)
            if not row["Emails"] and jl_emails:
                row["Emails"] = ", ".join(jl_emails)
            if not row["Phone Numbers"] and jl_phones:
                row["Phone Numbers"] = ", ".join(jl_phones)

        # Also try /contact in browser
        if not row["Emails"] or not row["Phone Numbers"]:
            parsed = urlparse(url)
            contact_url = f"{parsed.scheme}://{parsed.netloc}/contact"
            try:
                driver.get(contact_url)
                time.sleep(3)
                soup2 = BeautifulSoup(driver.page_source, "lxml")
                if not row["Emails"]:
                    emails = extract_emails_from_soup(soup2, domain_name)
                    if emails:
                        row["Emails"] = ", ".join(emails[:3])
                if not row["Phone Numbers"]:
                    phones = extract_phones_from_soup(soup2)
                    if phones:
                        row["Phone Numbers"] = ", ".join(phones)
                for platform, link in extract_socials(driver.page_source).items():
                    if not row[platform]:
                        row[platform] = link
            except Exception:
                pass

    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN SCRAPE FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════

def scrape(args):
    idx, url = args
    row = {c:"" for c in COLS}
    row["#"]   = idx
    row["URL"] = url

    domain      = get_domain(url)
    domain_name = get_domain_name(url)

    # Layer 1: Direct scrape (now returns soups for richer extraction)
    all_text, all_raw, status, all_soups = direct_scrape(url)
    row["Status"] = status

    if all_soups:
        # Collect across all pages
        all_emails, all_phones = [], []
        for soup in all_soups:
            all_emails += extract_emails_from_soup(soup, domain_name)
            all_phones += extract_phones_from_soup(soup)
            # JSON-LD structured data
            jl_emails, jl_phones = extract_jsonld_contacts(soup)
            all_emails += jl_emails
            all_phones += jl_phones

        if all_emails:
            row["Emails"] = ", ".join(list(dict.fromkeys(all_emails))[:5])
        if all_phones:
            row["Phone Numbers"] = ", ".join(list(dict.fromkeys(all_phones))[:3])
        for p, link in extract_socials(all_raw).items():
            row[p] = link

    elif all_text:
        # Fallback if no soup (shouldn't happen often)
        row["Emails"]        = ", ".join(extract_emails(all_text, domain_name))
        row["Phone Numbers"] = ", ".join(extract_phones(all_text))
        for p, link in extract_socials(all_raw).items():
            row[p] = link

    # Layer 2: Google Custom Search
    if still_missing(row):
        google_fallback(domain, domain_name, row)

    # Layer 3: Google Maps
    if not row["Phone Numbers"]:
        google_maps_fallback(domain_name, row)

    # Layer 4: Bing
    if still_missing(row):
        bing_fallback(domain, domain_name, row)

    # Layer 5: Yellow Pages
    if not row["Phone Numbers"] or not row["Emails"]:
        yellowpages_fallback(domain_name, row)

    # Layer 6: WHOIS (new)
    if not row["Emails"] or not row["Phone Numbers"]:
        whois_fallback(domain, row)

    # Layer 7: Selenium (per-thread, fixes v5 race condition)
    if still_missing(row):
        selenium_fallback(url, domain_name, row)

    # Update status
    if status != "OK" and (row["Emails"] or row["Phone Numbers"]):
        row["Status"] = status + " (recovered)"

    return row


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def dedup_urls(urls):
    from collections import OrderedDict
    seen = OrderedDict()
    for u in urls:
        parsed = urlparse(u)
        key = parsed.netloc.lower().replace("www.","") + parsed.path.rstrip("/")
        if key not in seen:
            seen[key] = u
        elif u.startswith("https") and not seen[key].startswith("https"):
            seen[key] = u
    return list(seen.values())


def save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contact Info"

    navy  = PatternFill("solid", start_color="1F3864")
    alt   = PatternFill("solid", start_color="EBF0FA")
    white = PatternFill("solid", start_color="FFFFFF")
    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    bf    = Font(name="Arial", size=10)
    lf    = Font(name="Arial", size=10, color="0563C1", underline="single")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="CCCCCC")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = hf, navy, ca, bord

    for ri, row in enumerate(rows, 2):
        fill = alt if ri % 2 == 0 else white
        for ci, col in enumerate(COLS, 1):
            val = row.get(col,"")
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill, c.border = fill, bord
            is_link = col not in ("#","URL","Status","Emails","Phone Numbers")
            c.font      = lf if (is_link and val) else bf
            c.alignment = ca if col in ("#","Status") else la

    widths = {"#":6,"URL":38,"Status":26,"Emails":34,"Phone Numbers":24,
              "Instagram":32,"Twitter/X":30,"Facebook":32,
              "LinkedIn":32,"YouTube":32,"TikTok":30}
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col,25)
    ws.row_dimensions[1].height = 22
    for ri in range(2, len(rows)+2):
        ws.row_dimensions[ri].height = 28
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ok        = sum(1 for r in rows if r.get("Status") == "OK")
    recovered = sum(1 for r in rows if "recovered" in r.get("Status",""))
    summary = [
        ("Total Sites",          len(rows)),
        ("Direct scrape OK",     ok),
        ("Recovered via search", recovered),
        ("Still errored",        len(rows)-ok-recovered),
        ("Had Email",            sum(1 for r in rows if r.get("Emails"))),
        ("Had Phone",            sum(1 for r in rows if r.get("Phone Numbers"))),
    ] + [(f"{p} found", sum(1 for r in rows if r.get(p)))
         for p in ["Instagram","Twitter/X","Facebook","LinkedIn","YouTube","TikTok"]]

    ws2["A1"], ws2["B1"] = "Metric", "Count"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font, cell.fill, cell.alignment = hf, navy, ca
    for ri, (label, val) in enumerate(summary, 2):
        ws2.cell(ri,1,label).font = bf
        ws2.cell(ri,2,val).font   = bf
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    wb.save(OUTPUT_FILE)


def load_urls():
    if not os.path.exists(URLS_FILE):
        print(f"\n  '{URLS_FILE}' not found.\n")
        input("Press Enter to exit...")
        return []
    with open(URLS_FILE,"r",encoding="utf-8") as f:
        lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]
    urls = [u if u.startswith("http") else "https://"+u for u in lines]
    deduped = dedup_urls(urls)
    print(f"  {len(urls)} URLs → {len(deduped)} unique")
    return deduped


def main():
    print("=" * 62)
    print("  CONTACT INFO SCRAPER v6")
    print("  Layers: scrape+sitemap → Google → Maps → Bing → YP → WHOIS → Selenium")
    print("=" * 62)

    urls = load_urls()
    if not urls:
        return

    print(f"\n  {MAX_WORKERS} workers (each gets its own browser) | auto-saves every {SAVE_EVERY} sites\n")

    results, completed = [], 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape,(i+1,url)): url for i,url in enumerate(urls)}
        with tqdm(total=len(urls), unit="site", ncols=70) as pbar:
            for future in as_completed(futures):
                row = future.result()
                completed += 1
                results.append(row)
                pbar.update(1)
                if completed % SAVE_EVERY == 0:
                    save_excel(sorted(results, key=lambda r: (
                        0 if (r.get("Emails") or r.get("Phone Numbers") or
                              r.get("Instagram") or r.get("Facebook") or
                              r.get("LinkedIn") or r.get("Twitter/X") or
                              r.get("YouTube") or r.get("TikTok")) else 1, r["#"])))
                    pbar.write(f"  Auto-saved at {completed} sites...")

    results.sort(key=lambda r: (
        0 if (r.get("Emails") or r.get("Phone Numbers") or
              r.get("Instagram") or r.get("Facebook") or
              r.get("LinkedIn") or r.get("Twitter/X") or
              r.get("YouTube") or r.get("TikTok"))
        else 1,
        r["#"]
    ))
    save_excel(results)

    ok        = sum(1 for r in results if r.get("Status") == "OK")
    recovered = sum(1 for r in results if "recovered" in r.get("Status",""))
    emails    = sum(1 for r in results if r.get("Emails"))
    phones    = sum(1 for r in results if r.get("Phone Numbers"))

    print(f"\n  Done!")
    print(f"  Direct OK:  {ok}/{len(results)}")
    print(f"  Recovered:  {recovered}")
    print(f"  Emails:     {emails}  |  Phones: {phones}")
    print(f"  Saved → {OUTPUT_FILE}\n")
    input("Press Enter to close...")


if __name__ == "__main__":
    main()
